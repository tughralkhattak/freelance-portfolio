"""
Sales Report Generator  —  portfolio demo
------------------------------------------
Reads a messy CSV of raw sales rows and produces a clean, formatted Excel
report with totals, a per-product summary, and a bar chart — in one click.

This is the kind of "turn boring manual work into one button" automation
clients pay $30-$100 for on Fiverr/Upwork.

HOW A CLIENT RUNS IT (no coding needed):
    1. Put their data in  sales_data.csv
    2. Double-click  run_report.bat
    3. Get  Sales_Report.xlsx

Requires: pandas, openpyxl   (already in the project venv)
"""

import csv
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Run:  pip install openpyxl   (then try again)")

HERE = Path(__file__).parent
SRC = HERE / "sales_data.csv"
OUT = HERE / "Sales_Report.xlsx"


def read_sales(path):
    """Read raw rows, skipping blanks and fixing common messiness."""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            product = (r.get("Product") or "").strip()
            if not product:
                continue  # skip empty rows
            try:
                qty = int(float(r.get("Quantity", 0) or 0))
                price = float(str(r.get("Price", 0)).replace("$", "").strip() or 0)
            except ValueError:
                continue  # skip rows with bad numbers
            rows.append({
                "Date": (r.get("Date") or "").strip(),
                "Product": product,
                "Quantity": qty,
                "Price": price,
                "Total": round(qty * price, 2),
            })
    return rows


def summarize(rows):
    """Total revenue per product."""
    by_product = defaultdict(float)
    for r in rows:
        by_product[r["Product"]] += r["Total"]
    return dict(sorted(by_product.items(), key=lambda kv: kv[1], reverse=True))


def build_excel(rows, summary):
    wb = Workbook()

    # ---- Sheet 1: clean detail table ----
    ws = wb.active
    ws.title = "Sales Detail"
    headers = ["Date", "Product", "Quantity", "Price", "Total"]
    head_fill = PatternFill("solid", fgColor="2E5A88")
    head_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r, row in enumerate(rows, 2):
        for c, key in enumerate(headers, 1):
            cell = ws.cell(r, c, row[key])
            cell.border = border
            if key in ("Price", "Total"):
                cell.number_format = '"$"#,##0.00'
            if key == "Quantity":
                cell.alignment = Alignment(horizontal="center")

    total_row = len(rows) + 2
    ws.cell(total_row, 4, "GRAND TOTAL").font = Font(bold=True)
    grand = ws.cell(total_row, 5, f"=SUM(E2:E{len(rows)+1})")
    grand.font = Font(bold=True)
    grand.number_format = '"$"#,##0.00'
    grand.fill = PatternFill("solid", fgColor="FCE8B2")

    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # ---- Sheet 2: summary + chart ----
    ws2 = wb.create_sheet("Summary")
    ws2.cell(1, 1, "Product").font = head_font
    ws2.cell(1, 2, "Revenue").font = head_font
    ws2.cell(1, 1).fill = head_fill
    ws2.cell(1, 2).fill = head_fill
    for i, (prod, rev) in enumerate(summary.items(), 2):
        ws2.cell(i, 1, prod)
        c = ws2.cell(i, 2, round(rev, 2))
        c.number_format = '"$"#,##0.00'
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 16

    chart = BarChart()
    chart.title = "Revenue by Product"
    chart.type = "bar"
    chart.style = 10
    n = len(summary)
    data = Reference(ws2, min_col=2, min_row=1, max_row=n + 1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=n + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws2.add_chart(chart, "D2")

    wb.save(OUT)


def main():
    if not SRC.exists():
        raise SystemExit(f"Put your data in {SRC.name} next to this script.")
    rows = read_sales(SRC)
    if not rows:
        raise SystemExit("No valid sales rows found in the CSV.")
    summary = summarize(rows)
    build_excel(rows, summary)
    grand = sum(r["Total"] for r in rows)
    print(f"Done. {len(rows)} rows processed, {len(summary)} products.")
    print(f"Total revenue: ${grand:,.2f}")
    print(f"Report saved: {OUT}")


if __name__ == "__main__":
    main()
